"""Bid templates: import PDF/Excel/CSV bid lists and map BOQ items onto them."""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.bid import BidTemplate, BidTemplateLine
from app.models.boq import BOQ
from app.services.csi_mapper import enrich_quantity_item, looks_like_csi, normalize_csi_code, normalize_unit


def list_templates(db: Session, project_id: int) -> list[BidTemplate]:
    return list(
        db.scalars(
            select(BidTemplate)
            .options(selectinload(BidTemplate.lines))
            .where(BidTemplate.project_id == project_id)
            .order_by(BidTemplate.created_at.desc())
        ).all()
    )


def get_active_template(db: Session, project_id: int) -> BidTemplate | None:
    return db.scalar(
        select(BidTemplate)
        .options(selectinload(BidTemplate.lines))
        .where(BidTemplate.project_id == project_id, BidTemplate.is_active.is_(True))
        .order_by(BidTemplate.created_at.desc())
    )


def set_active_template(db: Session, project_id: int, template_id: int) -> BidTemplate:
    templates = list(db.scalars(select(BidTemplate).where(BidTemplate.project_id == project_id)).all())
    found = None
    for t in templates:
        t.is_active = t.id == template_id
        if t.id == template_id:
            found = t
    if not found:
        raise ValueError("Bid template not found")
    db.commit()
    return db.scalar(
        select(BidTemplate).options(selectinload(BidTemplate.lines)).where(BidTemplate.id == template_id)
    )  # type: ignore[return-value]


def delete_template(db: Session, project_id: int, template_id: int) -> None:
    tmpl = db.scalar(select(BidTemplate).where(BidTemplate.id == template_id, BidTemplate.project_id == project_id))
    if not tmpl:
        raise ValueError("Bid template not found")
    db.delete(tmpl)
    db.commit()


def import_bid_template(
    db: Session,
    *,
    project_id: int,
    user_id: int,
    path: Path,
    filename: str,
    name: str | None = None,
) -> BidTemplate:
    rows = _parse_bid_file(path, filename)
    if not rows:
        raise ValueError("No bid line items found in file. Expect columns like CSI/Code, Description, Unit.")

    # Deactivate previous templates; new upload becomes active
    existing = list(db.scalars(select(BidTemplate).where(BidTemplate.project_id == project_id)).all())
    for t in existing:
        t.is_active = False

    template = BidTemplate(
        project_id=project_id,
        name=name or f"Bid list — {filename}",
        source_filename=filename,
        is_active=True,
        notes=f"Imported {len(rows)} line(s) from {filename}",
        created_by=user_id,
    )
    db.add(template)
    db.flush()

    for idx, row in enumerate(rows, start=1):
        csi = normalize_csi_code(row.get("csi_code") or row.get("item_code"))
        if csi and not looks_like_csi(csi):
            # keep as item_code only
            item_code = csi
            csi_val = None
        else:
            item_code = row.get("item_code") or csi
            csi_val = csi if looks_like_csi(csi) else None
        db.add(
            BidTemplateLine(
                template_id=template.id,
                line_number=str(row.get("line_number") or idx),
                csi_code=csi_val,
                item_code=item_code,
                description=str(row["description"]),
                unit=normalize_unit(row.get("unit")),
                default_rate=row.get("default_rate"),
                sort_order=idx,
            )
        )
    db.commit()
    return db.scalar(
        select(BidTemplate).options(selectinload(BidTemplate.lines)).where(BidTemplate.id == template.id)
    )  # type: ignore[return-value]


def map_boq_to_template(db: Session, *, boq_id: int, template_id: int | None = None) -> dict[str, Any]:
    boq = db.scalar(select(BOQ).options(selectinload(BOQ.items)).where(BOQ.id == boq_id))
    if not boq:
        raise ValueError("BOQ not found")

    if template_id:
        template = db.scalar(
            select(BidTemplate)
            .options(selectinload(BidTemplate.lines))
            .where(BidTemplate.id == template_id, BidTemplate.project_id == boq.project_id)
        )
    else:
        template = get_active_template(db, boq.project_id)
    if not template:
        raise ValueError("No active bid template. Upload a bid list (PDF/Excel/CSV) first.")

    lines = list(template.lines)
    matched = 0
    unmatched = 0
    details: list[dict[str, Any]] = []

    for item in boq.items:
        # Ensure CSI on BOQ item first
        enriched = enrich_quantity_item(
            {
                "description": item.description,
                "category": item.category,
                "unit": item.unit,
                "item_code": item.item_code,
                "csi_code": getattr(item, "csi_code", None),
                "confidence": float(item.confidence) if item.confidence is not None else None,
            }
        )
        item.csi_code = enriched.get("csi_code")
        if enriched.get("item_code") and not item.item_code:
            item.item_code = enriched["item_code"]
        item.unit = enriched.get("unit") or item.unit

        hit, score, method = _match_line(
            lines,
            description=item.description,
            unit=item.unit,
            csi_code=item.csi_code,
            item_code=item.item_code,
        )
        if hit:
            matched += 1
            item.bid_template_line_id = hit.id
            item.bid_match_confidence = score
            # Align to bidder's schedule (description, codes, unit)
            item.description = hit.description
            if hit.csi_code:
                item.csi_code = hit.csi_code
            if hit.item_code:
                item.item_code = hit.item_code
            item.unit = (hit.unit or item.unit or "UNIT").upper()
            if hit.default_rate is not None and item.rate is None:
                item.rate = Decimal(str(hit.default_rate))
                item.amount = Decimal(str(item.quantity)) * item.rate
            details.append(
                {
                    "boq_item_id": item.id,
                    "description": item.description,
                    "matched_line": hit.line_number,
                    "matched_description": hit.description,
                    "csi_code": item.csi_code,
                    "confidence": score,
                    "method": method,
                }
            )
        else:
            unmatched += 1
            item.bid_template_line_id = None
            item.bid_match_confidence = None
            item.unit = (item.unit or "UNIT").upper()
            details.append(
                {
                    "boq_item_id": item.id,
                    "description": item.description,
                    "matched_line": None,
                    "csi_code": item.csi_code,
                    "confidence": 0,
                    "method": "unmatched",
                }
            )

    db.commit()
    return {
        "template_id": template.id,
        "template_name": template.name,
        "boq_id": boq.id,
        "matched": matched,
        "unmatched": unmatched,
        "total": matched + unmatched,
        "details": details,
    }


def apply_template_to_items(items: list[dict[str, Any]], lines: list[BidTemplateLine]) -> list[dict[str, Any]]:
    """Legacy: enrich extracted items with matching template metadata (item-first)."""
    out: list[dict[str, Any]] = []
    for item in items:
        enriched = enrich_quantity_item(item)
        hit, score, method = _match_line(
            lines,
            description=str(enriched.get("description") or ""),
            unit=str(enriched.get("unit") or ""),
            csi_code=enriched.get("csi_code"),
            item_code=enriched.get("item_code"),
        )
        if hit:
            enriched["bid_template_line_id"] = hit.id
            enriched["bid_match_confidence"] = score
            enriched["bid_match_method"] = method
            if hit.csi_code:
                enriched["csi_code"] = hit.csi_code
            if hit.item_code:
                enriched["item_code"] = hit.item_code
            if hit.default_rate is not None and enriched.get("rate") is None:
                enriched["rate"] = hit.default_rate
        out.append(enriched)
    return out


def build_boq_items_from_template(
    extracted: list[dict[str, Any]],
    lines: list[BidTemplateLine],
) -> list[dict[str, Any]]:
    """Match plan takeoff to the bid template — only include items needed for this project.

    - Matched template lines (with evidence from plans/CAD) become BOQ rows using
      the template's Standard Bid Item Number, description, and unit.
    - Unmatched takeoff rows are kept as Unmapped takeoff for engineer review.
    - Template lines with no plan evidence are NOT dumped into the BOQ.
    """
    sorted_lines = sorted(lines, key=lambda line: (line.sort_order, line.id))
    enriched_extracted = [enrich_quantity_item(dict(item)) for item in extracted]

    line_hits: dict[int, list[tuple[dict[str, Any], float, str]]] = {line.id: [] for line in sorted_lines}
    unmatched: list[dict[str, Any]] = []

    for item in enriched_extracted:
        hit, score, method = _match_line(
            sorted_lines,
            description=str(item.get("description") or ""),
            unit=str(item.get("unit") or ""),
            csi_code=item.get("csi_code"),
            item_code=item.get("item_code"),
        )
        if hit and score > 0:
            line_hits[hit.id].append((item, score, method))
        else:
            unmatched.append(item)

    out: list[dict[str, Any]] = []
    for line in sorted_lines:
        hits = line_hits.get(line.id) or []
        if not hits:
            continue  # skip unused bid items — not needed for this project
        total_qty = 0.0
        for payload, _score, _method in hits:
            try:
                total_qty += float(payload.get("quantity") or 0)
            except (TypeError, ValueError):
                pass
        if total_qty <= 0:
            continue
        best_item, best_score, best_method = max(hits, key=lambda row: row[1])
        rate = line.default_rate if line.default_rate is not None else best_item.get("rate")
        out.append(
            {
                "item_code": line.item_code,
                "csi_code": line.csi_code,
                "description": line.description,
                "category": best_item.get("category") or "Bid schedule",
                "unit": (line.unit or best_item.get("unit") or "UNIT").upper(),
                "quantity": round(total_qty, 4),
                "rate": rate,
                "source_document_id": best_item.get("source_document_id"),
                "source_page": best_item.get("source_page"),
                "source_reference": best_item.get("source_reference")
                or f"Bid line {line.line_number}",
                "calculation_method": (
                    best_item.get("calculation_method")
                    or f"Matched to bid template line {line.line_number} ({best_method})"
                ),
                "confidence": best_item.get("confidence"),
                "bid_template_line_id": line.id,
                "bid_match_confidence": best_score,
                "bid_match_method": best_method,
            }
        )

    for item in unmatched:
        out.append(
            {
                **item,
                "unit": str(item.get("unit") or "UNIT").upper(),
                "category": item.get("category") or "Unmapped takeoff",
                "bid_template_line_id": None,
                "bid_match_confidence": 0,
                "bid_match_method": "unmapped",
                "calculation_method": (
                    (item.get("calculation_method") or "Takeoff")
                    + " — not matched to active bid template"
                ),
            }
        )
    return out


def _token_set(text: str) -> set[str]:
    stop = {"the", "and", "for", "with", "of", "a", "an", "to", "in", "on", "or"}
    tokens = re.findall(r"[a-z0-9.]+", text.lower())
    return {t for t in tokens if len(t) > 1 and t not in stop}


def _size_token(text: str) -> str | None:
    m = re.search(r"(\d{1,2}(?:\.\d+)?)\s*-?\s*(?:inch|in|\"|'')", text or "", re.I)
    if not m:
        m = re.search(r"\b(\d{1,2}(?:\.\d+)?)\s*\"", text or "")
    if not m:
        return None
    try:
        val = float(m.group(1))
    except ValueError:
        return None
    if abs(val - int(val)) < 0.01:
        return f"{int(val)}"
    return f"{val:g}"


def _units_compatible(a: str, b: str) -> bool:
    na, nb = normalize_unit(a), normalize_unit(b)
    if na == nb:
        return True
    # Only allow soft match when one side is unknown — never LF↔EA
    if na == "unit" or nb == "unit":
        return True
    return False


def _match_line(
    lines: list[BidTemplateLine],
    *,
    description: str,
    unit: str,
    csi_code: str | None,
    item_code: str | None,
) -> tuple[BidTemplateLine | None, float, str]:
    desc = description.lower().strip()
    unit_n = normalize_unit(unit)
    desc_tokens = _token_set(desc)
    desc_size = _size_token(description)

    if csi_code:
        for line in lines:
            if line.csi_code and normalize_csi_code(line.csi_code) == normalize_csi_code(csi_code):
                if _units_compatible(line.unit, unit_n):
                    return line, 96.0, "csi_code"
    if item_code:
        code = re.sub(r"\s+", "", item_code.lower().strip())
        for line in lines:
            for candidate in (line.item_code, line.csi_code):
                if candidate and re.sub(r"\s+", "", candidate.lower().strip()) == code:
                    if _units_compatible(line.unit, unit_n):
                        return line, 94.0, "item_code"

    for line in lines:
        if line.description.lower().strip() == desc and _units_compatible(line.unit, unit_n):
            return line, 92.0, "exact_description_unit"

    for line in lines:
        ld = line.description.lower().strip()
        if desc and ld and (desc in ld or ld in desc) and _units_compatible(line.unit, unit_n):
            line_size = _size_token(line.description)
            if desc_size and line_size and desc_size != line_size:
                continue
            return line, 82.0, "fuzzy_description"

    # Token overlap (e.g. "Aggregate Base Course" ↔ "Aggregate Base Course 3/4\")
    best: BidTemplateLine | None = None
    best_score = 0.0
    for line in lines:
        if not _units_compatible(line.unit, unit_n):
            continue
        line_size = _size_token(line.description)
        if desc_size and line_size and desc_size != line_size:
            continue
        line_tokens = _token_set(line.description)
        if not desc_tokens or not line_tokens:
            continue
        overlap = len(desc_tokens & line_tokens) / max(len(desc_tokens), 1)
        score = overlap * 100.0
        if desc_size and line_size and desc_size == line_size:
            score = min(100.0, score + 12.0)
        if score >= 65 and score > best_score:
            best = line
            best_score = score
    if best:
        return best, round(min(best_score, 88.0), 1), "token_overlap"

    return None, 0.0, "unmatched"


def _parse_bid_file(path: Path, filename: str) -> list[dict[str, Any]]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return _parse_excel(path)
    if ext == ".csv":
        return _parse_csv(path)
    if ext == ".pdf":
        return _parse_pdf(path)
    # try text extract for other docs
    if ext in {".txt"}:
        text = path.read_text(encoding="utf-8", errors="ignore")
        return _parse_free_text(text)
    raise ValueError("Bid template must be PDF, Excel (.xlsx/.xls), or CSV")


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    for i, r in enumerate(reader, start=1):
        mapped = _row_from_dict(r, fallback_line=i)
        if mapped:
            rows.append(mapped)
    return rows


def _parse_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    def find(*names: str) -> int | None:
        for name in names:
            for h, i in idx.items():
                if name == h or name in h:
                    return i
        return None

    line_i = find("line", "item no", "item#", "no.", "#")
    code_i = find("csi", "item code", "code", "bid item")
    desc_i = find("description", "desc", "item", "particular")
    unit_i = find("unit", "uom")
    rate_i = find("rate", "unit price", "price")
    if desc_i is None:
        desc_i = 0

    rows: list[dict[str, Any]] = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        values = list(row)
        if not values or desc_i >= len(values) or not values[desc_i]:
            continue
        code = str(values[code_i]).strip() if code_i is not None and code_i < len(values) and values[code_i] else None
        unit = str(values[unit_i]).strip() if unit_i is not None and unit_i < len(values) and values[unit_i] else "unit"
        rate = None
        if rate_i is not None and rate_i < len(values) and values[rate_i] is not None:
            try:
                rate = float(str(values[rate_i]).replace(",", ""))
            except ValueError:
                rate = None
        line_no = str(values[line_i]).strip() if line_i is not None and line_i < len(values) and values[line_i] else str(n)
        rows.append(
            {
                "line_number": line_no,
                "csi_code": code if code and looks_like_csi(code) else None,
                "item_code": code,
                "description": str(values[desc_i]).strip(),
                "unit": unit,
                "default_rate": rate,
            }
        )
    wb.close()
    return rows


def _parse_pdf(path: Path) -> list[dict[str, Any]]:
    from app.models.document import DocumentType
    from app.services.extractors import extract_file

    content = extract_file(path, DocumentType.PDF)
    rows: list[dict[str, Any]] = []
    for table in content.tables or []:
        table_rows = table.get("rows") or []
        if len(table_rows) < 2:
            continue
        header = [c.lower() for c in table_rows[0]]
        code_i = _find_col(header, ["csi", "code", "item code", "bid"])
        desc_i = _find_col(header, ["description", "desc", "item", "particular"])
        unit_i = _find_col(header, ["unit", "uom"])
        rate_i = _find_col(header, ["rate", "price", "unit price"])
        line_i = _find_col(header, ["line", "no", "#", "item no"])
        if desc_i is None:
            continue
        for n, row in enumerate(table_rows[1:], start=1):
            if desc_i >= len(row) or not str(row[desc_i]).strip():
                continue
            code = str(row[code_i]).strip() if code_i is not None and code_i < len(row) and row[code_i] else None
            unit = str(row[unit_i]).strip() if unit_i is not None and unit_i < len(row) and row[unit_i] else "unit"
            rate = None
            if rate_i is not None and rate_i < len(row) and row[rate_i]:
                try:
                    rate = float(re.sub(r"[^\d.]", "", str(row[rate_i])) or "0") or None
                except ValueError:
                    rate = None
            rows.append(
                {
                    "line_number": str(row[line_i]).strip() if line_i is not None and line_i < len(row) and row[line_i] else str(n),
                    "csi_code": code if code and looks_like_csi(code) else None,
                    "item_code": code,
                    "description": str(row[desc_i]).strip(),
                    "unit": unit,
                    "default_rate": rate,
                }
            )
    if rows:
        return rows
    return _parse_free_text(content.text or "")


def _parse_free_text(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for i, line in enumerate(text.splitlines(), start=1):
        line = line.strip()
        if len(line) < 8:
            continue
        # Pattern: CSI CODE  Description  UNIT  [rate]
        m = re.match(
            r"^(?:(\d+)\s+)?(\d{2}\s*\d{2}\s*\d{2}|\w[\w\-]{1,20})?\s*[-:]?\s*(.+?)\s+(m3|m2|m|cy|sy|lf|ea|nos|ls|t|kg)\b(?:\s+([\d,\.]+))?",
            line,
            flags=re.I,
        )
        if not m:
            continue
        code = (m.group(2) or "").strip() or None
        rows.append(
            {
                "line_number": m.group(1) or str(i),
                "csi_code": code if code and looks_like_csi(code) else None,
                "item_code": code,
                "description": m.group(3).strip(),
                "unit": m.group(4),
                "default_rate": float(m.group(5).replace(",", "")) if m.group(5) else None,
            }
        )
    return rows


def _row_from_dict(r: dict[str, Any], fallback_line: int) -> dict[str, Any] | None:
    lower = {str(k).lower().strip(): v for k, v in r.items() if k is not None}

    def get(*names: str) -> Any:
        for n in names:
            for k, v in lower.items():
                if n == k or n in k:
                    return v
        return None

    desc = get("description", "desc", "item", "particular")
    if not desc or not str(desc).strip():
        return None
    code = get("csi", "csi_code", "item_code", "code", "bid item")
    code_s = str(code).strip() if code else None
    rate_raw = get("rate", "unit price", "price")
    rate = None
    if rate_raw is not None and str(rate_raw).strip():
        try:
            rate = float(str(rate_raw).replace(",", ""))
        except ValueError:
            rate = None
    return {
        "line_number": str(get("line", "item no", "no") or fallback_line),
        "csi_code": code_s if code_s and looks_like_csi(code_s) else None,
        "item_code": code_s,
        "description": str(desc).strip(),
        "unit": str(get("unit", "uom") or "unit"),
        "default_rate": rate,
    }


def _find_col(header: list[str], names: list[str]) -> int | None:
    for i, h in enumerate(header):
        for n in names:
            if n == h or n in h:
                return i
    return None
