"""Build BOQ records from document analyses and export Excel."""

from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.models.analysis import DocumentAnalysis
from app.models.boq import BOQ, BOQItem, BOQItemStatus, BOQStatus
from app.models.cad import CadModel
from app.models.project import Project
from app.services.bid_service import build_boq_items_from_template, get_active_template
from app.services.boq_groups import assign_group_category, group_items
from app.services.csi_mapper import enrich_quantity_item
from app.services.processing import load_findings

# AutoVAD standard: confidence below this → Engineer Review
CONFIDENCE_VERIFIED_THRESHOLD = 97

BOQ_EXPORT_HEADERS = [
    "Item Number",
    "Standard Bid Item Number",
    "Item Description",
    "Unit",
    "Quantity",
    "Cost",
    "Total Cost",
    "AI Confidence",
    "Status",
    "Source",
    "Calculation Method",
]


def _money2(value: Decimal | float | None) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _qty2(value: Decimal | float | None) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def standard_bid_item_number(item: BOQItem) -> str:
    """Agency/state bid code from template; empty when AutoVAD default / unmapped."""
    if item.bid_template_line_id and item.item_code:
        return str(item.item_code)
    return ""


def status_label_for_item(item: BOQItem) -> str:
    """Display status for AutoVAD BOQ standard."""
    if item.status == BOQItemStatus.VERIFIED:
        return "Verified"
    if item.status in {BOQItemStatus.NEEDS_REVIEW, BOQItemStatus.DRAFT}:
        return "Engineer Review"
    if item.status == BOQItemStatus.APPROVED:
        return "Verified"
    return "Engineer Review"


def resolve_item_status(confidence: float | None, *, force_review: bool = False) -> BOQItemStatus:
    if force_review:
        return BOQItemStatus.NEEDS_REVIEW
    if confidence is not None and float(confidence) >= CONFIDENCE_VERIFIED_THRESHOLD:
        return BOQItemStatus.VERIFIED
    return BOQItemStatus.NEEDS_REVIEW


def _merge_item(merged: dict[str, dict], item: dict) -> None:
    item = enrich_quantity_item(item)
    code_key = (item.get("csi_code") or item.get("item_code") or "").strip().lower()
    key = f"{code_key}|{(item.get('description') or '').strip().lower()}|{(item.get('unit') or '').strip().lower()}"
    if not item.get("description"):
        return
    if key not in merged:
        merged[key] = dict(item)
        return
    existing = merged[key]
    try:
        existing_qty = Decimal(str(existing.get("quantity") or 0))
        new_qty = Decimal(str(item.get("quantity") or 0))
        existing_conf = Decimal(str(existing.get("confidence") or 0))
        new_conf = Decimal(str(item.get("confidence") or 0))
    except Exception:
        return
    if new_conf > existing_conf:
        merged[key] = dict(item)
    elif new_conf == existing_conf and new_qty > existing_qty:
        existing["quantity"] = float(new_qty)


def generate_boq_for_project(
    db: Session,
    project: Project,
    user_id: int,
    *,
    document_ids: list[int] | None = None,
) -> BOQ:
    """Build a BOQ from all analyses/CAD, or only the given document_ids (per-file BOQ)."""
    analyses = list(
        db.scalars(select(DocumentAnalysis).where(DocumentAnalysis.project_id == project.id)).all()
    )
    cad_models = list(db.scalars(select(CadModel).where(CadModel.project_id == project.id)).all())

    scope_ids = {int(i) for i in document_ids} if document_ids else None
    if scope_ids is not None:
        analyses = [a for a in analyses if a.document_id in scope_ids]
        cad_models = [c for c in cad_models if c.document_id in scope_ids]

    active = get_active_template(db, project.id)
    has_template = bool(active and active.lines)

    if not analyses and not cad_models:
        if scope_ids:
            raise ValueError(
                "No analysis for the selected file(s). Run Analyze (PDF) or Process CAD (DWG/DXF) on that file first."
            )
        raise ValueError("No analyzed documents or CAD models. Upload design plans and run Analyze first.")

    # Prefer CadModel quantities when both CAD analysis mirror and CadModel exist
    # for the same document (avoids near-duplicate BOQ lines).
    import json

    cad_doc_ids = {
        c.document_id
        for c in cad_models
        if c.quantities_json and str(c.quantities_json).strip() not in {"", "[]", "null"}
    }

    merged: dict[str, dict] = {}
    for analysis in analyses:
        if analysis.document_id in cad_doc_ids:
            continue
        findings = load_findings(analysis)
        for item in findings.get("items") or []:
            payload = dict(item)
            payload.setdefault("source_document_id", analysis.document_id)
            _merge_item(merged, payload)

    for cad in cad_models:
        if not cad.quantities_json:
            continue
        try:
            qty_items = json.loads(cad.quantities_json)
        except json.JSONDecodeError:
            qty_items = []
        for item in qty_items:
            payload = dict(item)
            payload["source_document_id"] = cad.document_id
            payload["calculation_method"] = payload.get("calculation_method") or "CAD geometry takeoff"
            _merge_item(merged, payload)

    extracted = list(merged.values())

    if not extracted:
        raise ValueError(
            "No quantities found from the selected file(s). "
            "Run Analyze / Process CAD first, then Generate BOQ again."
            if scope_ids
            else "No quantities found from design plans. Upload PDF/DWG and run Analyze / Process CAD first."
        )

    if has_template:
        # Only bid items evidenced in the plans, aligned to the active agency template.
        items_list = build_boq_items_from_template(extracted, list(active.lines))
        matched = sum(
            1
            for i in items_list
            if i.get("bid_template_line_id") and float(i.get("bid_match_confidence") or 0) > 0
        )
        unmapped = sum(1 for i in items_list if i.get("bid_match_method") == "unmapped")
        notes_extra = (
            f" Matched plan takeoff to bid template '{active.name}' "
            f"({matched} template item(s); {unmapped} unmapped takeoff item(s)). "
            "Unused bid-list lines were omitted."
        )
        if not items_list:
            raise ValueError(
                "Could not match any plan quantities to the active bid template. "
                "Re-analyze plans after uploading the template, or check descriptions/units."
            )
    else:
        # AutoVAD default: CSI-enriched takeoff BOQ
        items_list = [enrich_quantity_item(dict(item)) for item in extracted]
        notes_extra = (
            " Generated with AutoVAD default CSI schedule (no bid template uploaded)."
            " Upload a bid list so Generate BOQ maps only the bid items needed for this project."
        )

    from app.models.document import Document

    scope_label = ""
    if scope_ids:
        docs = list(db.scalars(select(Document).where(Document.id.in_(scope_ids))).all())
        names = [d.original_filename for d in docs]
        if len(names) == 1:
            scope_label = f" · {names[0]}"
        elif names:
            scope_label = f" · {len(names)} files"
        notes_extra += f" Scope: selected document_id(s) {sorted(scope_ids)}."

    latest_version = db.scalar(select(func.max(BOQ.version)).where(BOQ.project_id == project.id)) or 0
    boq = BOQ(
        project_id=project.id,
        title=f"{project.name} - BOQ v{latest_version + 1}{scope_label}",
        version=latest_version + 1,
        status=BOQStatus.AI_GENERATED,
        currency="USD",
        notes=(
            "Generated from document AI + CAD with CSI codes, units, and confidence scores."
            + notes_extra
        ),
        created_by=user_id,
    )
    db.add(boq)
    db.flush()

    # Persist in municipal group order with continuous Item No. 1, 2, 3, …
    ordered_items: list[dict] = []
    for _group_name, group_rows in group_items(
        items_list,
        get_description=lambda i: str(i.get("description") or ""),
        get_category=lambda i: i.get("category"),
    ):
        ordered_items.extend(group_rows)

    for idx, item in enumerate(ordered_items, start=1):
        conf = item.get("confidence")
        conf_f = float(conf) if conf is not None else None
        bid_match_conf = item.get("bid_match_confidence")
        try:
            bid_match_f = float(bid_match_conf) if bid_match_conf is not None else None
        except (TypeError, ValueError):
            bid_match_f = None
        # Ensure category is the EOQ group label
        grouped = assign_group_category(dict(item))
        force_review = (
            grouped.get("bid_match_method") in {"unmatched", "unmapped", "fuzzy_description"}
            or (bool(grouped.get("bid_template_line_id")) and float(grouped.get("quantity") or 0) == 0)
            or (
                bool(grouped.get("bid_template_line_id"))
                and bid_match_f is not None
                and bid_match_f < 85.0
            )
            or str(grouped.get("category") or "").lower() == "unmapped takeoff"
            or (
                not grouped.get("source_reference")
                and not grouped.get("source_document_id")
                and not grouped.get("calculation_method")
            )
        )
        status = resolve_item_status(conf_f, force_review=force_review)
        qty = _qty2(grouped.get("quantity"))
        rate = _money2(grouped["rate"]) if grouped.get("rate") is not None else None
        amount = _money2(qty * rate) if rate is not None else None
        unit = str(grouped.get("unit") or "UNIT").strip().upper() or "UNIT"
        db.add(
            BOQItem(
                boq_id=boq.id,
                item_number=str(idx),
                item_code=grouped.get("item_code"),
                csi_code=grouped.get("csi_code"),
                description=str(grouped.get("description")),
                category=grouped.get("category"),
                unit=unit,
                quantity=qty,
                rate=rate,
                amount=amount,
                source_document_id=grouped.get("source_document_id"),
                source_page=grouped.get("source_page"),
                source_reference=grouped.get("source_reference"),
                calculation_method=grouped.get("calculation_method"),
                confidence=Decimal(str(conf)) if conf is not None else None,
                bid_template_line_id=grouped.get("bid_template_line_id"),
                bid_match_confidence=grouped.get("bid_match_confidence"),
                status=status,
            )
        )

    db.commit()
    return db.scalar(
        select(BOQ).options(selectinload(BOQ.items)).where(BOQ.id == boq.id)
    )  # type: ignore[return-value]


def list_project_boqs(db: Session, project_id: int) -> list[BOQ]:
    return list(
        db.scalars(
            select(BOQ)
            .options(selectinload(BOQ.items))
            .where(BOQ.project_id == project_id)
            .order_by(BOQ.version.desc())
        ).all()
    )


def get_boq_item(db: Session, item_id: int) -> BOQItem | None:
    return db.scalar(select(BOQItem).where(BOQItem.id == item_id))


def update_boq_item(
    db: Session,
    item: BOQItem,
    *,
    status: BOQItemStatus | None = None,
    quantity: Decimal | None = None,
    description: str | None = None,
    unit: str | None = None,
    item_code: str | None = None,
    review_note: str | None = None,
) -> BOQItem:
    """Apply engineer review edits to one BOQ line."""
    if description is not None:
        cleaned = description.strip()
        if cleaned:
            item.description = cleaned
    if unit is not None:
        item.unit = unit.strip().upper() or item.unit
    if item_code is not None:
        item.item_code = item_code.strip() or None
    if quantity is not None:
        item.quantity = _qty2(quantity)
        if item.rate is not None:
            item.amount = _money2(item.quantity * item.rate)
    if status is not None:
        item.status = status
    if review_note:
        method = item.calculation_method or ""
        note = f"[review] {review_note.strip()}"
        item.calculation_method = f"{method} | {note}".strip(" |") if method else note
    db.commit()
    db.refresh(item)
    return item


def get_boq(db: Session, boq_id: int) -> BOQ | None:
    return db.scalar(select(BOQ).options(selectinload(BOQ.items)).where(BOQ.id == boq_id))


def _sorted_boq_items(boq: BOQ) -> list[BOQItem]:
    return sorted(
        boq.items,
        key=lambda x: int(x.item_number) if str(x.item_number).isdigit() else 0,
    )


def _grouped_boq_items(boq: BOQ) -> list[tuple[str, list[BOQItem]]]:
    """Return EOQ sections with items (item number order preserved within each group)."""
    items = _sorted_boq_items(boq)
    return group_items(
        items,
        get_description=lambda i: i.description,
        get_category=lambda i: i.category,
    )


def export_boq_csv(boq: BOQ) -> bytes:
    """Full BOQ columns + Group column, section separators, Item No. 1, 2, 3…"""
    buffer = StringIO()
    writer = csv.writer(buffer)

    # Keep Group as its own column so Excel/Sheets always show grouping clearly
    headers = [
        "Item Number",
        "Group",
        "Standard Bid Item Number",
        "Item Description",
        "Unit",
        "Quantity",
        "Cost",
        "Total Cost",
        "AI Confidence",
        "Status",
        "Source",
        "Calculation Method",
    ]
    writer.writerow(headers)

    serial = 1
    first_group = True
    for group_name, items in _grouped_boq_items(boq):
        if not first_group:
            writer.writerow([])  # blank line between groups
        first_group = False

        # Explicit section banner (shows even without using the Group column)
        writer.writerow(
            [
                "",
                group_name,
                "",
                f"===== {group_name} =====",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

        for item in items:
            qty = _qty2(item.quantity)
            cost = _money2(item.rate)
            total = _money2(qty * cost) if cost is not None else None
            writer.writerow(
                [
                    str(serial),
                    group_name,
                    standard_bid_item_number(item),
                    item.description,
                    (item.unit or "UNIT").upper(),
                    f"{qty:.2f}",
                    f"{cost:.2f}" if cost is not None else "",
                    f"{total:.2f}" if total is not None else "",
                    f"{float(item.confidence):.2f}" if item.confidence is not None else "",
                    status_label_for_item(item),
                    item.source_reference or "",
                    item.calculation_method or "",
                ]
            )
            serial += 1

    # UTF-8 BOM so Excel on Windows detects encoding / columns correctly
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def export_boq_excel(boq: BOQ) -> bytes:
    """Full AutoVAD BOQ columns with municipal-style section grouping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    headers = BOQ_EXPORT_HEADERS
    col_count = len(headers)

    title_font = Font(name="Calibri", size=14, bold=True, color="D9FF43")
    header_fill = PatternFill("solid", fgColor="0D1F19")
    header_font = Font(color="D9FF43", bold=True)
    section_fill = PatternFill("solid", fgColor="D9E2F3")
    section_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(1, 1, "ESTIMATE OF QUANTITIES")
    title_cell.font = title_font
    title_cell.fill = PatternFill("solid", fgColor="0D1F19")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    row_idx = 3
    item_rows: list[int] = []
    serial = 1

    for group_name, items in _grouped_boq_items(boq):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_count)
        sec = ws.cell(row_idx, 1, group_name)
        sec.font = section_font
        sec.fill = section_fill
        sec.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, col_count + 1):
            cell = ws.cell(row_idx, c)
            cell.fill = section_fill
            cell.border = border
        row_idx += 1

        for item in items:
            qty = float(_qty2(item.quantity))
            cost = _money2(item.rate)
            status = status_label_for_item(item)

            ws.cell(row_idx, 1, serial).alignment = Alignment(horizontal="center")
            ws.cell(row_idx, 2, standard_bid_item_number(item)).alignment = Alignment(horizontal="center")
            ws.cell(row_idx, 3, item.description).alignment = Alignment(horizontal="left", wrap_text=True)
            ws.cell(row_idx, 4, (item.unit or "UNIT").upper()).alignment = Alignment(horizontal="center")

            qty_cell = ws.cell(row_idx, 5, qty)
            qty_cell.number_format = "0.00"
            qty_cell.alignment = Alignment(horizontal="right")

            if cost is not None:
                cost_cell = ws.cell(row_idx, 6, float(cost))
                cost_cell.number_format = "0.00"
                total_cell = ws.cell(row_idx, 7, f"=E{row_idx}*F{row_idx}")
                total_cell.number_format = "0.00"
            else:
                ws.cell(row_idx, 6, "")
                ws.cell(row_idx, 7, "")

            if item.confidence is not None:
                conf_cell = ws.cell(row_idx, 8, round(float(item.confidence), 2))
                conf_cell.number_format = "0.00"
            else:
                ws.cell(row_idx, 8, "")

            status_cell = ws.cell(row_idx, 9, status)
            if status == "Verified":
                status_cell.font = Font(color="006100", bold=True)
                status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                status_cell.font = Font(color="9C0006", bold=True)
                status_cell.fill = PatternFill("solid", fgColor="FFC7CE")

            ws.cell(row_idx, 10, item.source_reference or "")
            ws.cell(row_idx, 11, item.calculation_method or "")

            for c in range(1, col_count + 1):
                ws.cell(row_idx, c).border = border

            item_rows.append(row_idx)
            serial += 1
            row_idx += 1

    last_row = max(3, row_idx - 1)
    if item_rows:
        status_dv = DataValidation(
            type="list",
            formula1='"Verified,Engineer Review"',
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid status",
            error="Choose Verified or Engineer Review",
        )
        status_dv.add(f"I3:I{last_row}")
        ws.add_data_validation(status_dv)

        green_fill = PatternFill("solid", fgColor="C6EFCE")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        red_font = Font(color="9C0006", bold=True)
        ws.conditional_formatting.add(
            f"I3:I{last_row}",
            FormulaRule(formula=['$I3="Verified"'], fill=green_fill, font=green_font),
        )
        ws.conditional_formatting.add(
            f"I3:I{last_row}",
            FormulaRule(formula=['$I3="Engineer Review"'], fill=red_fill, font=red_font),
        )

    widths = [12, 22, 42, 10, 12, 12, 12, 14, 16, 28, 28]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A3"

    meta = wb.create_sheet("Meta")
    meta.append(["BOQ Title", boq.title])
    meta.append(["Version", boq.version])
    meta.append(["Status", boq.status.value])
    meta.append(["Currency", boq.currency])
    meta.append(["Notes", boq.notes or ""])
    meta.append(["Generated by", "AutoVAD"])
    meta.append(
        [
            "Layout",
            "Full BOQ columns with municipal section grouping "
            "(Removals, Grading, Watermain, Sanitary Sewer, …)",
        ]
    )
    meta.append(
        [
            "Template rule",
            "User bid template when active; otherwise AutoVAD default CSI schedule",
        ]
    )
    meta.append(
        [
            "Status rule",
            f"Verified when AI confidence >= {CONFIDENCE_VERIFIED_THRESHOLD}; else Engineer Review",
        ]
    )
    meta.append(["Columns", ", ".join(BOQ_EXPORT_HEADERS)])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
