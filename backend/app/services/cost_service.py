from __future__ import annotations

import csv
import io
import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session, selectinload

from app.models.boq import BOQ
from app.models.cost import CostEstimate, SORItem


def clear_sor(db: Session, project_id: int) -> None:
    db.execute(delete(SORItem).where(SORItem.project_id == project_id))
    db.commit()


def import_sor_file(db: Session, project_id: int, path: Path, filename: str) -> list[SORItem]:
    clear_sor(db, project_id)
    rows: list[tuple[str | None, str, str, Decimal]] = []
    ext = path.suffix.lower()

    if ext == ".csv":
        text = path.read_text(encoding="utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            desc = (r.get("description") or r.get("Description") or r.get("item") or r.get("Item") or "").strip()
            unit = (r.get("unit") or r.get("Unit") or "unit").strip()
            rate_raw = r.get("rate") or r.get("Rate") or r.get("amount") or "0"
            code = (r.get("item_code") or r.get("code") or r.get("Code") or None)
            if not desc:
                continue
            rows.append((code, desc, unit, Decimal(str(rate_raw).replace(",", "") or "0")))
    else:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}
        desc_i = next((idx[h] for h in idx if "desc" in h or h == "item"), 0)
        unit_i = next((idx[h] for h in idx if "unit" in h), 1)
        rate_i = next((idx[h] for h in idx if "rate" in h), 2)
        code_i = next((idx[h] for h in idx if "code" in h), None)
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row)
            if not values or desc_i >= len(values) or not values[desc_i]:
                continue
            code = str(values[code_i]).strip() if code_i is not None and code_i < len(values) and values[code_i] else None
            desc = str(values[desc_i]).strip()
            unit = str(values[unit_i]).strip() if unit_i < len(values) and values[unit_i] else "unit"
            rate = Decimal(str(values[rate_i] or 0).replace(",", ""))
            rows.append((code, desc, unit, rate))
        wb.close()

    items: list[SORItem] = []
    for code, desc, unit, rate in rows:
        item = SORItem(project_id=project_id, item_code=code, description=desc, unit=unit, rate=rate)
        db.add(item)
        items.append(item)
    db.commit()
    return items


def list_sor(db: Session, project_id: int) -> list[SORItem]:
    return list(db.scalars(select(SORItem).where(SORItem.project_id == project_id).order_by(SORItem.id.asc())).all())


def _match_rate(sor_items: list[SORItem], description: str, unit: str, item_code: str | None) -> SORItem | None:
    desc = description.lower().strip()
    unit_n = unit.lower().strip()
    if item_code:
        for s in sor_items:
            if s.item_code and s.item_code.lower() == item_code.lower():
                return s
    for s in sor_items:
        if s.description.lower().strip() == desc and s.unit.lower().strip() == unit_n:
            return s
    for s in sor_items:
        if desc in s.description.lower() or s.description.lower() in desc:
            if s.unit.lower().strip() == unit_n or True:
                return s
    return None


def generate_cost_estimate(db: Session, *, project_id: int, boq_id: int, user_id: int) -> CostEstimate:
    boq = db.scalar(select(BOQ).options(selectinload(BOQ.items)).where(BOQ.id == boq_id, BOQ.project_id == project_id))
    if not boq:
        raise ValueError("BOQ not found for project")
    sor_items = list_sor(db, project_id)
    if not sor_items:
        raise ValueError("Upload a Schedule of Rates (SOR) first")

    breakdown = []
    total = Decimal("0")
    category_totals: dict[str, float] = {}

    for item in boq.items:
        match = _match_rate(sor_items, item.description, item.unit, item.item_code)
        rate = match.rate if match else None
        amount = (item.quantity * rate) if rate is not None else None
        if amount is not None:
            total += amount
            cat = item.category or "General"
            category_totals[cat] = category_totals.get(cat, 0) + float(amount)
        breakdown.append(
            {
                "boq_item_id": item.id,
                "description": item.description,
                "category": item.category,
                "unit": item.unit,
                "quantity": float(item.quantity),
                "rate": float(rate) if rate is not None else None,
                "amount": float(amount) if amount is not None else None,
                "matched_sor": match.description if match else None,
                "matched": match is not None,
            }
        )
        if rate is not None:
            item.rate = rate
            item.amount = amount

    estimate = CostEstimate(
        project_id=project_id,
        boq_id=boq_id,
        title=f"Cost estimate for {boq.title}",
        currency=boq.currency,
        total_amount=total,
        breakdown_json=json.dumps(
            {"items": breakdown, "category_totals": category_totals},
            ensure_ascii=True,
        ),
        created_by=user_id,
    )
    db.add(estimate)
    db.commit()
    db.refresh(estimate)
    return estimate
