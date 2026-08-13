"""Raw text / table extraction from uploaded civil documents."""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import load_workbook

from app.models.document import DocumentType


@dataclass
class PageText:
    page: int
    text: str


@dataclass
class ExtractedContent:
    text: str
    page_count: int | None = None
    pages: list[PageText] = field(default_factory=list)
    tables: list[dict] = field(default_factory=list)


def extract_file(path: Path, document_type: DocumentType) -> ExtractedContent:
    if document_type == DocumentType.PDF:
        return _extract_pdf(path)
    if document_type == DocumentType.EXCEL:
        return _extract_excel(path)
    if document_type == DocumentType.CSV:
        return _extract_csv(path)
    if document_type == DocumentType.IMAGE:
        return ExtractedContent(
            text=f"[Image file: {path.name}] OCR/vision extraction will use AI when OpenAI key is configured.",
            page_count=1,
            pages=[PageText(page=1, text=f"Image document: {path.name}")],
        )
    if document_type == DocumentType.ZIP:
        return _extract_zip(path)
    return ExtractedContent(text=f"Unsupported or unknown file type: {path.name}")


def _extract_pdf(path: Path) -> ExtractedContent:
    pages: list[PageText] = []
    tables: list[dict] = []

    with fitz.open(path) as doc:
        for i, page in enumerate(doc, start=1):
            text = page.get_text("text") or ""
            pages.append(PageText(page=i, text=text))

    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                for table in page.extract_tables() or []:
                    if not table:
                        continue
                    tables.append({"page": i, "rows": table[:80]})
    except Exception:
        pass

    full_text = "\n\n".join(f"--- Page {p.page} ---\n{p.text}" for p in pages if p.text.strip())
    return ExtractedContent(text=full_text, page_count=len(pages), pages=pages, tables=tables)


def _extract_excel(path: Path) -> ExtractedContent:
    wb = load_workbook(path, data_only=True, read_only=True)
    chunks: list[str] = []
    tables: list[dict] = []
    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True, max_row=200, max_col=20):
            values = [("" if c is None else str(c)).strip() for c in row]
            if any(values):
                rows.append(values)
        if rows:
            tables.append({"page": 1, "sheet": sheet.title, "rows": rows[:80]})
            lines = ["\t".join(r) for r in rows[:80]]
            chunks.append(f"--- Sheet {sheet.title} ---\n" + "\n".join(lines))
    wb.close()
    return ExtractedContent(text="\n\n".join(chunks), page_count=1, pages=[PageText(page=1, text="\n\n".join(chunks))], tables=tables)


def _extract_csv(path: Path) -> ExtractedContent:
    text = path.read_text(encoding="utf-8", errors="ignore")
    rows: list[list[str]] = []
    reader = csv.reader(io.StringIO(text))
    for i, row in enumerate(reader):
        if i >= 200:
            break
        rows.append([c.strip() for c in row])
    return ExtractedContent(
        text=text[:50000],
        page_count=1,
        pages=[PageText(page=1, text=text[:20000])],
        tables=[{"page": 1, "rows": rows[:80]}],
    )


def _extract_zip(path: Path) -> ExtractedContent:
    names: list[str] = []
    with zipfile.ZipFile(path, "r") as zf:
        names = zf.namelist()[:200]
    listing = "\n".join(names)
    text = f"ZIP archive contains {len(names)} entries:\n{listing}"
    return ExtractedContent(text=text, page_count=1, pages=[PageText(page=1, text=text)])
