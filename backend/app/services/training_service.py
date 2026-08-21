"""Training lab service: gold cases, AutoVAD runs, compare reports for agent training."""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.document import DocumentType
from app.models.training import (
    TrainingCase,
    TrainingCaseStatus,
    TrainingReport,
    TrainingRun,
    TrainingRunStatus,
)
from app.services.eoq_eval import compare_eoq
from app.services.documents import detect_document_type, guess_content_type
from app.services.storage import storage_service

logger = logging.getLogger(__name__)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def list_cases(db: Session) -> list[TrainingCase]:
    return list(db.scalars(select(TrainingCase).order_by(TrainingCase.updated_at.desc())).all())


def get_case(db: Session, case_id: int) -> TrainingCase | None:
    return db.scalar(
        select(TrainingCase)
        .options(selectinload(TrainingCase.runs).selectinload(TrainingRun.report))
        .where(TrainingCase.id == case_id)
    )


def create_case(
    db: Session,
    *,
    user_id: int,
    name: str,
    description: str | None = None,
    notes: str | None = None,
) -> TrainingCase:
    case = TrainingCase(
        name=name.strip() or "Untitled training case",
        description=description,
        notes=notes,
        status=TrainingCaseStatus.DRAFT,
        created_by=user_id,
    )
    db.add(case)
    db.commit()
    db.refresh(case)
    return case


def update_case(
    db: Session,
    case: TrainingCase,
    *,
    name: str | None = None,
    description: str | None = None,
    notes: str | None = None,
    status: str | None = None,
) -> TrainingCase:
    if name is not None:
        case.name = name.strip() or case.name
    if description is not None:
        case.description = description
    if notes is not None:
        case.notes = notes
    if status is not None:
        case.status = TrainingCaseStatus(status)
    case.updated_at = _now()
    db.commit()
    db.refresh(case)
    return case


def delete_case(db: Session, case: TrainingCase) -> None:
    for key in (case.sample_storage_key, case.expected_storage_key):
        if key:
            try:
                storage_service.delete_file(key)
            except Exception:
                pass
    db.delete(case)
    db.commit()


async def save_sample_file(db: Session, case: TrainingCase, *, filename: str, data: bytes) -> TrainingCase:
    if case.sample_storage_key:
        try:
            storage_service.delete_file(case.sample_storage_key)
        except Exception:
            pass
    safe = Path(filename).name.replace(" ", "_")
    key = f"training/{case.id}/{safe}"
    await storage_service.save_file(key, data)
    case.sample_filename = filename
    case.sample_storage_key = key
    case.sample_content_type = guess_content_type(filename)
    case.sample_file_size = len(data)
    # New sample invalidates prior AutoVAD EOQ / status
    case.actual_json = None
    case.actual_engine = None
    case.actual_notes = None
    case.analyzed_at = None
    case.updated_at = _now()
    _refresh_ready_status(case)
    db.commit()
    db.refresh(case)
    return case


def set_expected_json(db: Session, case: TrainingCase, payload: dict[str, Any] | list[Any]) -> TrainingCase:
    if isinstance(payload, list):
        payload = {"items": payload}
    if not isinstance(payload, dict) or not payload.get("items"):
        raise ValueError("Expected items must include an 'items' array with description/unit/quantity")
    case.expected_json = json.dumps(payload, ensure_ascii=False)
    case.updated_at = _now()
    _refresh_ready_status(case)
    db.commit()
    db.refresh(case)
    return case


async def save_expected_file(db: Session, case: TrainingCase, *, filename: str, data: bytes) -> TrainingCase:
    """Upload original Estimate Of Quantities (PDF / Excel / CSV / image / JSON) and parse gold items."""
    if case.expected_storage_key:
        try:
            storage_service.delete_file(case.expected_storage_key)
        except Exception:
            pass

    safe = Path(filename).name.replace(" ", "_")
    key = f"training/{case.id}/expected_{safe}"
    await storage_service.save_file(key, data)
    path = storage_service.resolve_local_path(key)

    items = parse_expected_eoq_file(path, filename)
    if not items:
        raise ValueError(
            "No Estimate Of Quantities items found in that file. "
            "For CAD/plan-sheet EOQ PDFs (tables drawn as graphics), ensure OPENAI_API_KEY is set "
            "so vision can read the sheet. Or upload Excel/CSV with Description, Unit, Quantity columns."
        )

    case.expected_filename = filename
    case.expected_storage_key = key
    case.expected_json = json.dumps(
        {
            "id": f"case_{case.id}",
            "name": case.name,
            "source_filename": filename,
            "items": items,
        },
        ensure_ascii=False,
    )
    case.updated_at = _now()
    _refresh_ready_status(case)
    db.commit()
    db.refresh(case)
    return case


def parse_expected_eoq_file(path: Path, filename: str) -> list[dict[str, Any]]:
    """Parse original EOQ from PDF, Excel, CSV, image, or JSON into gold items."""
    ext = Path(filename).suffix.lower() or path.suffix.lower()

    if ext == ".json":
        payload = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
        if isinstance(payload, list):
            rows = payload
        elif isinstance(payload, dict):
            rows = payload.get("items") or []
        else:
            rows = []
        return [_normalize_expected_item(r) for r in rows if _normalize_expected_item(r)]

    if ext in {".xlsx", ".xls"}:
        return _parse_expected_excel(path)
    if ext == ".csv":
        return _parse_expected_csv(path)
    if ext == ".pdf":
        return _parse_expected_pdf(path, filename)
    if ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp"}:
        return _parse_expected_image(path, filename)

    raise ValueError(
        "Original Estimate Of Quantities must be PDF, Excel (.xlsx/.xls), CSV, or image "
        "(PNG/JPG/TIF). JSON is also accepted for advanced/gold-set use."
    )


def _normalize_expected_item(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row or not isinstance(row, dict):
        return None
    desc = str(
        row.get("description")
        or row.get("desc")
        or row.get("item_description")
        or row.get("item description")
        or row.get("item")
        or ""
    ).strip()
    # Skip section headers / title noise
    if not desc:
        return None
    lower = desc.lower()
    if lower in {
        "item description",
        "description",
        "estimate of quantities",
        "for bidding purposes only",
    }:
        return None
    if re.fullmatch(r"(general|removals?|grading|surfacing|watermain|storm sewer|erosion|traffic).*", lower):
        # keep only if it also looks like a real bid line (has unit/qty elsewhere)
        if row.get("quantity") in (None, "", "-") and not (row.get("unit") or row.get("uom")):
            return None
    unit = str(row.get("unit") or row.get("uom") or "UNIT").strip() or "UNIT"
    qty_raw = row.get("quantity")
    if qty_raw is None:
        qty_raw = (
            row.get("qty")
            or row.get("approx")
            or row.get("approx. quantity")
            or row.get("approx quantity")
            or row.get("approximate quantity")
        )
    quantity: float | None = None
    if qty_raw is not None and str(qty_raw).strip() != "":
        try:
            quantity = float(re.sub(r"[^\d.\-]", "", str(qty_raw).replace(",", "")) or "nan")
            if quantity != quantity:  # NaN
                quantity = None
        except ValueError:
            quantity = None
    item_code = (
        row.get("item_code")
        or row.get("code")
        or row.get("std_bid_no")
        or row.get("std bid no")
        or row.get("bid_no")
    )
    item_no = row.get("item_no") or row.get("item no") or row.get("item#") or row.get("no")
    return {
        "description": desc,
        "unit": unit,
        "quantity": quantity,
        "category": row.get("category"),
        "item_code": str(item_code).strip() if item_code not in (None, "") else None,
        "item_no": str(item_no).strip() if item_no not in (None, "") else None,
        "required": bool(row.get("required", True)),
        "quantity_abs_tolerance": float(row.get("quantity_abs_tolerance", 1.0)),
        "quantity_tolerance": float(row.get("quantity_tolerance", 0.05)),
    }


def _parse_expected_excel(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        wb.close()
        return []
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_row]
    idx = {h: i for i, h in enumerate(headers)}

    def find(*names: str) -> int | None:
        for name in names:
            for h, i in idx.items():
                if name == h or name in h:
                    return i
        return None

    desc_i = find("description", "desc", "item description", "particular", "item")
    unit_i = find("unit", "uom")
    qty_i = find("quantity", "qty", "approx", "approx. quantity", "approx quantity")
    code_i = find("item code", "bid item", "std bid", "code", "csi")
    cat_i = find("category", "group", "division")
    if desc_i is None:
        desc_i = 0

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row)
        if not values or desc_i >= len(values) or not values[desc_i]:
            continue
        payload = {
            "description": values[desc_i],
            "unit": values[unit_i] if unit_i is not None and unit_i < len(values) else "UNIT",
            "quantity": values[qty_i] if qty_i is not None and qty_i < len(values) else None,
            "item_code": values[code_i] if code_i is not None and code_i < len(values) else None,
            "category": values[cat_i] if cat_i is not None and cat_i < len(values) else None,
        }
        item = _normalize_expected_item(payload)
        if item:
            rows.append(item)
    wb.close()
    return rows


def _parse_expected_csv(path: Path) -> list[dict[str, Any]]:
    import csv
    import io

    text = path.read_text(encoding="utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    for r in reader:
        item = _normalize_expected_item({str(k).lower(): v for k, v in r.items() if k is not None})
        if item:
            rows.append(item)
    return rows


def _parse_expected_pdf(path: Path, filename: str) -> list[dict[str, Any]]:
    from app.models.document import DocumentType
    from app.services.extractors import extract_file

    content = extract_file(path, DocumentType.PDF)
    rows: list[dict[str, Any]] = []
    for table in content.tables or []:
        table_rows = table.get("rows") or []
        if len(table_rows) < 2:
            continue
        header = [str(c).lower() for c in table_rows[0]]
        desc_i = _find_header_col(
            header,
            ["item description", "description", "desc", "particular", "item"],
        )
        unit_i = _find_header_col(header, ["unit", "uom"])
        qty_i = _find_header_col(header, ["approx. quantity", "approx quantity", "quantity", "qty", "approx"])
        code_i = _find_header_col(header, ["std bid no", "std bid", "bid no", "code", "csi", "item no", "item#"])
        if desc_i is None:
            continue
        # Heuristic: CAD title-block "tables" are usually 1–2 junk rows
        if len(table_rows) < 3:
            continue
        for row in table_rows[1:]:
            if desc_i >= len(row) or not str(row[desc_i]).strip():
                continue
            item = _normalize_expected_item(
                {
                    "description": row[desc_i],
                    "unit": row[unit_i] if unit_i is not None and unit_i < len(row) else "UNIT",
                    "quantity": row[qty_i] if qty_i is not None and qty_i < len(row) else None,
                    "item_code": row[code_i] if code_i is not None and code_i < len(row) else None,
                }
            )
            if item:
                rows.append(item)
    if rows:
        return rows

    # Text fallback (works for selectable-text EOQ PDFs with real line items in text)
    text = (content.text or "").strip()
    text_looks_like_eoq = bool(
        re.search(
            r"(item\s*description|approx\.?\s*quantity|std\s*bid|estimate of quantities|eoq)",
            text,
            re.I,
        )
    )
    if text_looks_like_eoq and len(text) > 400:
        ai_rows = _extract_expected_via_ai(text=text[:120000], filename=filename, images=None)
        if ai_rows:
            return ai_rows

    # CAD / plan-sheet EOQs (like City of Sioux Falls sheets) store the table as drawings —
    # text extractors only see the title block. Render pages and use vision.
    vision_rows = _extract_expected_pdf_via_vision(path, filename)
    if vision_rows:
        return vision_rows
    return []


def _render_pdf_pages_png(path: Path, *, max_pages: int = 6, zoom: float = 2.0) -> list[dict[str, str]]:
    """Rasterize PDF pages for vision EOQ extraction."""
    import base64

    import fitz

    doc = fitz.open(path)
    images: list[dict[str, str]] = []
    try:
        page_count = min(doc.page_count, max_pages)
        for i in range(page_count):
            pix = doc[i].get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            images.append(
                {
                    "page": str(i + 1),
                    "png_b64": base64.b64encode(pix.tobytes("png")).decode("ascii"),
                }
            )
    finally:
        doc.close()
    return images


def _extract_expected_pdf_via_vision(path: Path, filename: str) -> list[dict[str, Any]]:
    try:
        images = _render_pdf_pages_png(path)
    except Exception:
        return []
    if not images:
        return []
    # ask_openai_vision_json expects page as int-ish in dict; keep page key as int
    vision_images = [{"page": int(img["page"]), "png_b64": img["png_b64"]} for img in images]
    return _extract_expected_via_ai(text=None, filename=filename, images=vision_images)


def _parse_expected_image(path: Path, filename: str) -> list[dict[str, Any]]:
    import base64

    png_bytes: bytes
    try:
        import fitz

        doc = fitz.open(path)
        try:
            page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            png_bytes = pix.tobytes("png")
        finally:
            doc.close()
    except Exception:
        png_bytes = path.read_bytes()

    b64 = base64.b64encode(png_bytes).decode("ascii")
    return _extract_expected_via_ai(
        text=None,
        filename=filename,
        images=[{"png_b64": b64}],
    )


def _find_header_col(header: list[str], names: list[str]) -> int | None:
    for name in names:
        for i, h in enumerate(header):
            if name == h or name in h:
                return i
    return None


def _extract_expected_via_ai(
    *,
    text: str | None,
    filename: str,
    images: list[dict[str, str]] | None,
) -> list[dict[str, Any]]:
    from app.services.openai_client import ask_openai_json, ask_openai_vision_json, openai_configured

    if not openai_configured():
        return []

    system = (
        "Extract Estimate Of Quantities / EOQ line items from civil plan sheets. "
        "Sheets often use two side-by-side tables with columns: ITEM NO, STD BID NO, "
        "ITEM DESCRIPTION, UNIT, APPROX. QUANTITY. Read LEFT table fully, then RIGHT table. "
        "Ignore title blocks, logos, revision stamps, and category headers "
        "(e.g. Removals, Grading, Surfacing, Watermain) unless they are real bid lines. "
        "Return JSON only: {\"items\":[{\"description\":\"...\",\"unit\":\"LS|Ft|Each|Ton|CuYd|SqYd|...\", "
        "\"quantity\":number|null,\"item_code\":string|null,\"item_no\":string|null,\"category\":string|null}]}. "
        "item_code = STD BID NO when present. Include every measurable line."
    )
    user = (
        f"Original EOQ file: {filename}\n"
        "Extract all Estimate Of Quantities line items from every table on every page."
    )
    try:
        if images:
            vision_images = []
            for i, img in enumerate(images):
                page = img.get("page", i + 1)
                try:
                    page = int(page)
                except (TypeError, ValueError):
                    page = i + 1
                vision_images.append({"page": page, "png_b64": img["png_b64"]})
            data = ask_openai_vision_json(system=system, user=user, images=vision_images)
        else:
            data = ask_openai_json(system=system, user=f"{user}\n\nDocument text:\n{text or ''}")
        raw_items = data.get("items") if isinstance(data, dict) else None
        out: list[dict[str, Any]] = []
        for row in raw_items or []:
            item = _normalize_expected_item(row if isinstance(row, dict) else None)
            if item:
                out.append(item)
        return out
    except Exception:
        logger.exception("EOQ vision/text extraction failed for %s", filename)
        return []


def set_bid_catalog_json(db: Session, case: TrainingCase, payload: list[dict[str, Any]] | dict[str, Any]) -> TrainingCase:
    if isinstance(payload, dict):
        rows = payload.get("items") or payload.get("lines") or []
    else:
        rows = payload
    case.bid_catalog_json = json.dumps(list(rows), ensure_ascii=False)
    case.updated_at = _now()
    db.commit()
    db.refresh(case)
    return case


def _refresh_ready_status(case: TrainingCase) -> None:
    """Stage-aware status: draft → analyzed → ready."""
    if case.status == TrainingCaseStatus.ARCHIVED:
        return
    has_autovad = bool(case.actual_json)
    has_original = bool(case.expected_json)
    if has_autovad and has_original:
        case.status = TrainingCaseStatus.READY
    elif has_autovad:
        case.status = TrainingCaseStatus.ANALYZED
    else:
        case.status = TrainingCaseStatus.DRAFT


def _load_json(raw: str | None) -> Any:
    if not raw:
        return None
    return json.loads(raw)


def run_autovad_analyze(db: Session, case: TrainingCase) -> TrainingCase:
    """Stage 1: Analyze sample plan and store AutoVAD Estimate Of Quantities on the case."""
    if not case.sample_storage_key:
        raise ValueError("Upload a sample PDF/DWG/DXF file first")

    path = storage_service.resolve_local_path(case.sample_storage_key)
    if not path.exists():
        raise FileNotFoundError("Sample file missing from storage")

    doc_type = detect_document_type(case.sample_filename or path.name)
    bid_catalog = _load_json(case.bid_catalog_json) or []
    if isinstance(bid_catalog, dict):
        bid_catalog = bid_catalog.get("items") or bid_catalog.get("lines") or []

    actual_items, engine, notes = _run_autovad_takeoff(
        path=path,
        filename=case.sample_filename or path.name,
        document_type=doc_type,
        bid_catalog=list(bid_catalog),
    )
    if not actual_items:
        detail = (notes or "").strip()
        if "needs Autodesk" in detail or "AUTODESK_CLIENT" in detail or engine.startswith("autodesk_aps_pending"):
            raise ValueError(detail or "DWG requires Autodesk APS credentials.")
        raise ValueError(
            "AutoVAD produced 0 Estimate Of Quantities items. "
            + (f"{detail} " if detail else "")
            + "For DWG: ensure APS credentials are set (same as user projects), wait for APS translation to finish, "
            "or export DWG→DXF. For PDF: check OpenAI key."
        )

    case.actual_json = json.dumps(actual_items, ensure_ascii=False)
    case.actual_engine = engine
    case.actual_notes = notes
    case.analyzed_at = _now()
    case.updated_at = _now()
    _refresh_ready_status(case)
    db.commit()
    db.refresh(case)
    return case


def run_evaluation(db: Session, case: TrainingCase, *, user_id: int) -> TrainingRun:
    """Stage 3: Compare AutoVAD EOQ vs original EOQ and write a training report."""
    actual_items = _load_json(case.actual_json)
    if not isinstance(actual_items, list) or not actual_items:
        raise ValueError("Stage 1 incomplete — run Analyze on the sample plan first")
    if not case.expected_json:
        raise ValueError("Stage 2 incomplete — upload the original Estimate Of Quantities file first")

    run = TrainingRun(
        case_id=case.id,
        status=TrainingRunStatus.RUNNING,
        started_at=_now(),
        created_by=user_id,
        engine=case.actual_engine,
        actual_json=case.actual_json,
        analysis_notes=case.actual_notes,
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    try:
        expected = _load_json(case.expected_json)
        compare = compare_eoq(expected, actual_items)
        metrics = compare.to_dict()
        guidance, ai_generated = _build_training_guidance(
            case_name=case.name,
            filename=case.sample_filename or "",
            metrics=metrics,
            expected=expected,
            actual=actual_items,
        )

        run.status = TrainingRunStatus.COMPLETED
        run.finished_at = _now()
        run.error_message = None

        visual = metrics.get("visual") or {}
        report = TrainingReport(
            run_id=run.id,
            metrics_json=json.dumps(metrics, ensure_ascii=False),
            diffs_json=json.dumps(
                {
                    "misses": metrics.get("misses") or [],
                    "extras": metrics.get("extras") or [],
                    "qty_errors": metrics.get("qty_errors") or [],
                    "hits": metrics.get("hits") or [],
                    "visual": visual,
                    "near_misses": visual.get("near_misses") or [],
                    "matched": visual.get("matched") or [],
                    "line_audit": visual.get("line_audit") or [],
                    "by_category": visual.get("by_category") or [],
                    "summary": visual.get("summary") or {},
                },
                ensure_ascii=False,
            ),
            training_guidance=guidance,
            recall=str(metrics.get("recall")),
            precision_proxy=str(metrics.get("precision_proxy")),
            ai_generated=ai_generated,
        )
        db.add(report)
        case.updated_at = _now()
        db.commit()
        db.refresh(run)
        return get_run(db, run.id) or run
    except Exception as exc:
        run.status = TrainingRunStatus.FAILED
        run.error_message = str(exc)
        run.finished_at = _now()
        db.commit()
        db.refresh(run)
        raise


def run_training_case(db: Session, case: TrainingCase, *, user_id: int) -> TrainingRun:
    """Backward-compatible: analyze if needed, then evaluate."""
    if not case.actual_json:
        run_autovad_analyze(db, case)
        db.refresh(case)
    return run_evaluation(db, case, user_id=user_id)

def get_run(db: Session, run_id: int) -> TrainingRun | None:
    return db.scalar(
        select(TrainingRun)
        .options(selectinload(TrainingRun.report), selectinload(TrainingRun.case))
        .where(TrainingRun.id == run_id)
    )


def list_runs(db: Session, case_id: int) -> list[TrainingRun]:
    return list(
        db.scalars(
            select(TrainingRun)
            .options(selectinload(TrainingRun.report))
            .where(TrainingRun.case_id == case_id)
            .order_by(TrainingRun.id.desc())
        ).all()
    )


def _run_autovad_takeoff(
    *,
    path: Path,
    filename: str,
    document_type: DocumentType,
    bid_catalog: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str, str]:
    """Execute current AutoVAD extractors (PDF AI / CAD) without a project wrapper."""
    suffix = path.suffix.lower()

    # CAD path
    if suffix in {".dxf", ".dwg", ".xml", ".landxml", ".json"} or document_type in {
        DocumentType.DXF,
        DocumentType.DWG,
        DocumentType.LANDXML,
        DocumentType.CIVIL3D,
    }:
        return _run_cad_takeoff(path, filename)

    from app.services.ai_analysis import analyze_content
    from app.services.extractors import extract_file

    content = extract_file(path, document_type)
    result = analyze_content(
        filename=filename,
        content=content,
        document_id=0,
        bid_catalog=bid_catalog,
        file_path=path,
    )
    items = list(result.get("items") or [])
    engine = str(result.get("engine") or "heuristic")
    notes = str(result.get("notes") or result.get("summary") or "")
    if result.get("vision_coverage"):
        notes = (notes + f" | vision_coverage={result['vision_coverage']}").strip(" |")
    return items, engine, notes


def _run_cad_takeoff(path: Path, filename: str) -> tuple[list[dict[str, Any]], str, str]:
    """CAD takeoff for Training Lab — same engines as user projects (DXF local, DWG via APS)."""
    from app.config import get_settings
    from app.services.cad.aps_client import aps_configured, aps_status
    from app.services.cad.autodesk_aps import parse_civil3d_package, parse_dwg
    from app.services.cad.dxf_parser import parse_dxf
    from app.services.cad.landxml_parser import parse_landxml
    from app.services.cad.quantity_engine import build_quantities
    from app.services.openai_client import enrich_cad_quantities_with_openai, openai_configured

    suffix = path.suffix.lower()
    settings = get_settings()

    try:
        if suffix == ".dxf":
            extraction = parse_dxf(path)
            engine = "cad_dxf"
        elif suffix in {".xml", ".landxml"}:
            extraction = parse_landxml(path)
            engine = "cad_landxml"
        elif suffix == ".json":
            extraction = parse_civil3d_package(path)
            engine = str(extraction.get("engine") or "cad_json")
        elif suffix == ".dwg":
            # Same path as user-account Process CAD / Analyze for DWG
            extraction = parse_dwg(path)
            engine = str(extraction.get("engine") or "autodesk_aps")
        else:
            extraction = parse_dwg(path)
            engine = str(extraction.get("engine") or "cad_unknown")

        status = extraction.get("status")
        if status == "needs_autodesk":
            aps = aps_status()
            raise ValueError(
                "DWG processing needs Autodesk Platform Services (same as user projects). "
                "Set AUTODESK_CLIENT_ID and AUTODESK_CLIENT_SECRET in backend/.env, enable "
                "Design Automation + Model Derivative on the APS app, restart the backend, "
                "then try Analyze again. "
                f"APS configured={aps.get('configured')}. "
                "Or export DWG → DXF for local parsing without APS."
            )
        if status == "failed":
            raise ValueError(
                extraction.get("error")
                or extraction.get("summary")
                or "CAD/DWG processing failed via Autodesk APS."
            )

        items = build_quantities(
            extraction if isinstance(extraction, dict) else {},
            source_label=filename,
        )
        for hint in extraction.get("quantities_hint") or []:
            if hint.get("description") and hint.get("quantity") is not None:
                items.append(dict(hint))

        if settings.cad_openai_enrichment and openai_configured() and items:
            items = enrich_cad_quantities_with_openai(
                filename=filename,
                extraction_summary=str(extraction.get("summary") or ""),
                stats=extraction.get("stats") or {},
                quantities=items,
                layers=extraction.get("layers") or [],
                blocks=extraction.get("blocks") or [],
                pipes=extraction.get("pipes") or [],
                texts=extraction.get("texts") or [],
            )
            engine = f"{engine}+openai"

        from app.services.traffic_control import consolidate_traffic_control_signs

        items, tc_meta = consolidate_traffic_control_signs(items, allow_online_refresh=True)

        notes = str(extraction.get("summary") or f"CAD takeoff from {filename}")
        if tc_meta.get("sign_rows"):
            notes = (
                f"{notes} | Rolled {tc_meta['sign_rows']} traffic sign(s) into "
                f"Traffic Control ({tc_meta.get('total_sqft')} SqFt)."
            ).strip()
        if not items:
            # APS ran but quantity engine found nothing useful
            hint = (
                " APS is configured." if aps_configured() else " APS may be missing credentials."
            )
            notes = (
                f"{notes} No quantity candidates extracted from CAD geometry.{hint} "
                "Check layers/blocks or try DXF/LandXML export."
            )
        return list(items or []), engine, notes
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"CAD training takeoff failed: {exc}") from exc


def _build_training_guidance(
    *,
    case_name: str,
    filename: str,
    metrics: dict[str, Any],
    expected: Any,
    actual: list[dict[str, Any]],
) -> tuple[str, bool]:
    """Heuristic report always; enrich with OpenAI when configured."""
    base = _heuristic_guidance(case_name, filename, metrics)
    try:
        from app.services.openai_client import ask_openai_text, openai_configured

        if not openai_configured():
            return base, False
        system = (
            "You are an AutoVAD training coach for civil quantity takeoff agents. "
            "Given expected (gold) items vs AutoVAD actual items and a diff summary, "
            "write a concise training report for engineers who will fine-tune agents. "
            "Focus on: systematic mistakes, missing categories, size/unit errors, "
            "false extras, and concrete prompt/data fixes. Use markdown. No fluff."
        )
        user = json.dumps(
            {
                "case": case_name,
                "file": filename,
                "metrics": {
                    "recall": metrics.get("recall"),
                    "precision_proxy": metrics.get("precision_proxy"),
                    "expected_count": metrics.get("expected_count"),
                    "actual_count": metrics.get("actual_count"),
                    "misses_by_category": metrics.get("misses_by_category"),
                    "miss_count": len(metrics.get("misses") or []),
                    "extra_count": len(metrics.get("extras") or []),
                    "qty_error_count": len(metrics.get("qty_errors") or []),
                },
                "sample_misses": (metrics.get("misses") or [])[:12],
                "sample_extras": (metrics.get("extras") or [])[:12],
                "sample_qty_errors": (metrics.get("qty_errors") or [])[:12],
            },
            ensure_ascii=False,
        )
        ai = ask_openai_text(system=system, user=user, temperature=0.2)
        return f"{base}\n\n---\n\n## AI training coach\n\n{ai.strip()}", True
    except Exception:
        return base, False


def _heuristic_guidance(case_name: str, filename: str, metrics: dict[str, Any]) -> str:
    misses = metrics.get("misses") or []
    extras = metrics.get("extras") or []
    qty_errors = metrics.get("qty_errors") or []
    by_cat = metrics.get("misses_by_category") or {}
    lines = [
        f"# Training report — {case_name}",
        f"Source file: `{filename}`",
        "",
        "## Scores",
        f"- Recall: **{metrics.get('recall')}** (found expected items)",
        f"- Precision proxy: **{metrics.get('precision_proxy')}**",
        f"- Expected: {metrics.get('expected_count')} · AutoVAD: {metrics.get('actual_count')}",
        f"- Misses: {len(misses)} · Extras: {len(extras)} · Qty errors: {len(qty_errors)}",
        "",
        "## What to train on",
    ]
    if by_cat:
        lines.append("### Missing by category")
        for cat, n in sorted(by_cat.items(), key=lambda x: -x[1]):
            lines.append(f"- **{cat}**: {n} missing — boost extractor coverage for this group")
    if misses:
        lines.append("### Missed gold items (AutoVAD failed to find)")
        for m in misses[:15]:
            desc = m.get("description") if isinstance(m, dict) else str(m)
            unit = m.get("unit", "") if isinstance(m, dict) else ""
            lines.append(f"- `{desc}` ({unit}) — add similar examples to vision/text gold set")
    if extras:
        lines.append("### Extra AutoVAD items (not in gold)")
        for e in extras[:15]:
            lines.append(
                f"- `{e.get('description')}` ({e.get('unit')}) qty={e.get('quantity')} — "
                "reduce false positives or tighten filters"
            )
    if qty_errors:
        lines.append("### Quantity mismatches")
        for q in qty_errors[:15]:
            lines.append(f"- {q}")
    if not misses and not extras and not qty_errors:
        lines.append("- No major diffs — keep this case as a regression lock.")
    lines.extend(
        [
            "",
            "## Suggested next actions",
            "1. Add miss examples into agent gold prompts / fine-tune dataset.",
            "2. If extras dominate, tighten Vision/Label rules for this plan type.",
            "3. Re-run after prompt/model change and compare recall/precision.",
            "4. Export this report with the case when fine-tuning Bid Matcher or Vision agents.",
        ]
    )
    return "\n".join(lines)


def case_to_dict(case: TrainingCase, *, include_runs: bool = False) -> dict[str, Any]:
    expected = _load_json(case.expected_json)
    bid = _load_json(case.bid_catalog_json)
    actual = _load_json(case.actual_json)
    actual_items = actual if isinstance(actual, list) else []
    data: dict[str, Any] = {
        "id": case.id,
        "name": case.name,
        "description": case.description,
        "status": case.status.value if hasattr(case.status, "value") else case.status,
        "sample_filename": case.sample_filename,
        "sample_file_size": case.sample_file_size,
        "sample_content_type": case.sample_content_type,
        "has_sample": bool(case.sample_storage_key),
        "has_autovad_eoq": bool(actual_items),
        "autovad_item_count": len(actual_items),
        "actual_engine": case.actual_engine,
        "actual_notes": case.actual_notes,
        "analyzed_at": case.analyzed_at.isoformat() if case.analyzed_at else None,
        "has_expected": bool(case.expected_json),
        "expected_filename": case.expected_filename,
        "expected_item_count": len((expected or {}).get("items") or []) if isinstance(expected, dict) else 0,
        "can_evaluate": bool(actual_items) and bool(case.expected_json),
        "bid_catalog_count": len(bid) if isinstance(bid, list) else 0,
        "notes": case.notes,
        "created_by": case.created_by,
        "created_at": case.created_at.isoformat() if case.created_at else None,
        "updated_at": case.updated_at.isoformat() if case.updated_at else None,
    }
    if include_runs:
        data["runs"] = [run_to_dict(r) for r in (case.runs or [])]
        data["expected"] = expected
        data["bid_catalog"] = bid
        data["autovad_items"] = actual_items
    return data


def run_to_dict(run: TrainingRun) -> dict[str, Any]:
    actual = _load_json(run.actual_json) or []
    report = None
    if run.report:
        report = {
            "id": run.report.id,
            "metrics": _load_json(run.report.metrics_json),
            "diffs": _load_json(run.report.diffs_json),
            "training_guidance": run.report.training_guidance,
            "recall": run.report.recall,
            "precision_proxy": run.report.precision_proxy,
            "ai_generated": run.report.ai_generated,
            "created_at": run.report.created_at.isoformat() if run.report.created_at else None,
        }
    return {
        "id": run.id,
        "case_id": run.case_id,
        "status": run.status.value if hasattr(run.status, "value") else run.status,
        "engine": run.engine,
        "actual_item_count": len(actual) if isinstance(actual, list) else 0,
        "actual_items": actual if isinstance(actual, list) else [],
        "analysis_notes": run.analysis_notes,
        "error_message": run.error_message,
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "finished_at": run.finished_at.isoformat() if run.finished_at else None,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "report": report,
    }
