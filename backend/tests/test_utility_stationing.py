"""Tests for underground utility stationing / Excel detail tables."""

from app.services.cad.utility_stationing import (
    build_utilities_detail,
    feet_to_station,
    station_to_feet,
)


def test_station_roundtrip():
    assert station_to_feet("10+50") == 1050.0
    assert feet_to_station(1050) == "10+50"


def test_text_station_segments():
    extraction = {
        "texts": [
            {"layer": "NOTES", "text": '8" WATER MAIN STA 10+00 TO 12+50'},
            {"layer": "NOTES", "text": "12\" SANITARY SEWER STA 5+00 TO 7+25"},
        ],
        "polylines": [],
        "pipes": [],
        "blocks": [],
        "alignments": [],
        "lines": [],
    }
    detail = build_utilities_detail(extraction)
    assert detail["summary"]["segment_count"] >= 2
    utilities = {s["utility"] for s in detail["segments"]}
    assert "Water" in utilities
    assert "Sanitary Sewer" in utilities
    water = next(s for s in detail["segments"] if s["utility"] == "Water")
    assert water["quantity_lf"] == 250.0
    assert "Increasing" in water["direction"]


def test_geometry_projection_and_bend():
    # Alignment along X axis 0→500
    alignment_pts = [[0, 0], [500, 0]]
    # Water pipe parallel at y=10 from x=100 to x=300, with a bend
    pipe_pts = [[100, 10], [200, 10], [200, 40], [300, 40]]
    extraction = {
        "texts": [],
        "alignments": [{"name": "CL Alignment", "points": alignment_pts, "length": 500, "sta_start": "0+00"}],
        "polylines": [
            {"layer": "WATER_MAIN", "name": "8in WM", "points": pipe_pts, "length": 0},
        ],
        "pipes": [],
        "blocks": [
            {
                "name": "BEND_8",
                "layer": "WATER_FITTINGS",
                "type": "Bend",
                "insert": [200, 10],
                "size": '8"',
            }
        ],
        "lines": [],
    }
    detail = build_utilities_detail(extraction)
    assert detail["summary"]["segment_count"] >= 1
    seg = detail["segments"][0]
    assert seg["from_station"]
    assert seg["to_station"]
    assert seg["side_of_alignment"] in {"Left", "Right", "On", ""}
    assert detail["summary"]["connection_count"] >= 1
    conn = detail["connections"][0]
    assert conn["station"]


def test_excel_sheets_created():
    from openpyxl import load_workbook
    from io import BytesIO

    from app.models.eoq import EOQ, EOQItem, EOQItemStatus, EOQStatus
    from app.services.eoq_service import export_eoq_excel

    eoq = EOQ(
        id=1,
        project_id=1,
        title="Test EOQ",
        version=1,
        status=EOQStatus.AI_GENERATED,
        currency="USD",
        notes="",
    )
    eoq.items = [
        EOQItem(
            id=1,
            eoq_id=1,
            item_number="1",
            description="8-Inch Water Main",
            unit="LF",
            quantity=250,
            status=EOQItemStatus.NEEDS_REVIEW,
            confidence=90,
        )
    ]
    detail = {
        "segments": [
            {
                "utility": "Water",
                "size": "8-Inch",
                "description": "8-Inch Water",
                "from_station": "10+00",
                "to_station": "12+50",
                "direction": "Increasing station",
                "side_of_alignment": "Left",
                "quantity_lf": 250,
                "layer": "WATER",
                "alignment": "CL",
                "source": "test",
                "method": "test",
            }
        ],
        "connections": [
            {
                "utility": "Water",
                "connection_type": "Bend / Elbow",
                "size": "8-Inch",
                "station": "11+00",
                "direction_from_alignment": "Left",
                "offset_ft": "5.00",
                "quantity": 1,
                "unit": "EA",
                "layer": "FITTINGS",
                "alignment": "CL",
                "source": "test",
                "method": "test",
            }
        ],
    }
    content = export_eoq_excel(eoq, utilities_detail=detail)
    wb = load_workbook(BytesIO(content))
    assert "Estimate Of Quantities" in wb.sheetnames
    assert "Utility Stationing" in wb.sheetnames
    assert "Utility Connections" in wb.sheetnames
    assert wb["Utility Stationing"]["A2"].value == "Water"
    assert wb["Utility Connections"]["D2"].value == "11+00"
